#!/usr/bin/env python
"Print stock prices of companies I am interested in"

import yfinance as yf
import pandas as pd
import argparse
import os
import logging
from tabulate import tabulate
from openpyxl import load_workbook

# Setup logger
logger = logging.getLogger("PortfolioLogger")
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter("[%(asctime)s] %(levelname)s - %(message)s", datefmt='%Y-%m-%d %H:%M:%S')
handler.setFormatter(formatter)
logger.addHandler(handler)

# Fields to collect
FIELDS = [
    "Ticker",
    "Beta",
    "Opening Price",
    "Current Price",
    "52 Week High",
    "50 Day MA",
    "200 Day MA",
    "Indicator"
]

def fetch_stock_info(ticker):
    """
    Fetch relevant data for a given stock ticker.
    """
    try:
        #logger.info(f"{ticker}...")
        logger.info(f"\033[92m{ticker}\033[0m")
        stock = yf.Ticker(ticker)
        info = stock.info
        data = {
            "Ticker": ticker,
            "Beta": info.get("beta"),
            "Opening Price": info.get("regularMarketOpen"),
            "Current Price": info.get("regularMarketPrice"),
            "52 Week High": info.get("fiftyTwoWeekHigh"),
            "50 Day MA": info.get("fiftyDayAverage"),
            "200 Day MA": info.get("twoHundredDayAverage"),
        }
        return data
    except Exception as e:
        logger.error(f"Error fetching data for {ticker}: {e}")
        return {
            "Ticker": ticker,
            "Beta": None,
            "Opening Price": None,
            "Current Price": None,
            "52 Week High": None,
            "50 Day MA": None,
            "200 Day MA": None
        }

def write_values_preserving_formatting(input_file, updated_df, sheet_name="Ticker"):
    wb = load_workbook(input_file)
    ws = wb[sheet_name]

    # Map headers to column numbers
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map = {col_name: idx + 1 for idx, col_name in enumerate(header_row)}

    for i, row in updated_df.iterrows():
        excel_row = i + 2  # +2 because Excel is 1-indexed and row 1 is header
        for col in updated_df.columns:
            if col in col_map:
                ws.cell(row=excel_row, column=col_map[col], value=row[col])

    wb.save(input_file)
    logger.info(f"Formatting preserved while updating sheet '{sheet_name}' in file: {input_file}")

def process_excel(input_file):
    SHEET_NAME = "Tickers"

    if not os.path.isfile(input_file):
        logger.error(f"File '{input_file}' not found.")
        raise FileNotFoundError(f"File '{input_file}' not found.")

    logger.info(f"Reading sheet '{SHEET_NAME}' from file: {input_file}")
    excel_file = pd.ExcelFile(input_file)
    
    if SHEET_NAME not in excel_file.sheet_names:
        logger.error(f"Sheet named '{SHEET_NAME}' not found in the file.")
        raise ValueError(f"Sheet named '{SHEET_NAME}' not found.")

    df = pd.read_excel(excel_file, sheet_name=SHEET_NAME)

    if "Ticker" not in df.columns:
        logger.error("Missing required column: 'Ticker'")
        raise ValueError("Sheet must contain a 'Ticker' column.")

    logger.info("Fetching data for each ticker...")
    stock_data_list = [fetch_stock_info(ticker) for ticker in df["Ticker"]]

    stock_data_df = pd.DataFrame(stock_data_list).set_index("Ticker")
    df.set_index("Ticker", inplace=True)

    for col in FIELDS[1:]:
        if col == "Indicator" or col == "Beta":
            continue
        else:
            df[col] = stock_data_df[col]

    # take the new beta value only if it is not NaN
    df.update(stock_data_df[["Beta"]])

    df["Indicator"] = df.apply(
        lambda row: "Bull" if pd.notnull(row["50 Day MA"]) and pd.notnull(row["200 Day MA"]) and row["50 Day MA"] > row["200 Day MA"] else "Bear",
        axis=1
    )
    updated_df = df.reset_index()

    # Print updated values as a table
    logger.info("\n" + tabulate(updated_df[FIELDS], headers="keys", tablefmt="github", floatfmt=".2f"))

    # Write values back while preserving formatting
    write_values_preserving_formatting(input_file, updated_df, sheet_name=SHEET_NAME)

def main():
    parser = argparse.ArgumentParser(description="Fetch stock stats from Yahoo Finance and update Excel.")
    parser.add_argument("--input-file", required=True, help="Path to the input Excel file")
    
    args = parser.parse_args()
    process_excel(args.input_file)

if __name__ == "__main__":
    main()
